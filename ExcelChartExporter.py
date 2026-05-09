# -*- coding: utf-8 -*-
"""
Excel Chart Exporter for pyADR v3.5 (clean template edition)

Strategy: load `pyADR_excel_template.xlsx` (a 6-sheet workbook with
pre-built charts), overwrite the input cells of each sheet with
pyADR-computed values, save as user's output.

Template sheets and their input cells (rows 2..16, 15 steps):
  - 'output data'      A1  = "Sample ID = ..." stamp
  - 'normal isochron'  A:D = 39Ar/36Ar, σ, 40Ar/36Ar, σ
  - 'inverse isochron' A:D = 39Ar/40Ar, σ, 36Ar/40Ar, σ
  - 'Age spectrum'     A:C = Age (Ma), σ, cum%39Ar
  - 'Ca-K ratio'       A:C = Ca/K, σ, cum%39Ar
  - 'Cl-K ratio'       A:C = Cl/K, σ, cum%39Ar

All other cells (D..T) are formulas relative to A..C/D and re-evaluate
when Excel opens the workbook → charts re-draw automatically.

Author: pyADR Development Team
Date: 2026-05-07
"""

import os
import shutil
import numpy as np
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill


DEFAULT_TEMPLATE = 'pyADR_excel_template.xlsx'
TEMPLATE_MAX_STEPS = 15


def _resolve_template(explicit_path):
    if explicit_path and os.path.isfile(explicit_path):
        return explicit_path
    here = os.path.dirname(os.path.abspath(__file__))
    candidates = [
        os.path.join(here, DEFAULT_TEMPLATE),
        os.path.join(os.getcwd(), DEFAULT_TEMPLATE),
    ]
    for p in candidates:
        if os.path.isfile(p):
            return p
    return None


# ────────────────────────────────────────────────────────────────────────
class ExcelChartExporter:
    """Export pyADR step-heating data via the 6-sheet Excel template."""

    def __init__(self, data_file, mask, constants,
                 out_filename='pyADR_Export.xlsx',
                 work_dir=None, stats=None, template_path=None):
        self.data_file = data_file
        self.mask = np.asarray(mask, dtype=float)
        self.constants = constants
        self.out_filename = out_filename
        self.work_dir = work_dir if work_dir else os.path.dirname(data_file)
        self.stats = stats or {}
        self.template_path = _resolve_template(template_path)
        self.df = None
        self.sample_info = {}

    # ------------------------------------------------------------------
    def read_data(self):
        df = pd.read_csv(self.data_file)
        if 'Age(Ma)' in df.columns:
            df = df[df['Age(Ma)'].notna()].reset_index(drop=True)
        self.df = df
        for col in ('Samp#', 'Min', 'IRR'):
            if col in df.columns and len(df) > 0:
                self.sample_info[col] = df[col].iloc[0]
        for col in ('J', 'J_std'):
            if col in df.columns and len(df) > 0:
                try:
                    self.sample_info[col] = float(df[col].iloc[0])
                except Exception:
                    self.sample_info[col] = df[col].iloc[0]
        print(f"[ExcelExport] Read {len(df)} valid rows from {os.path.basename(self.data_file)}")

    def _cumulative_percent_array(self):
        col = '39Ar(k)(%)(step heating)'
        if col in self.df.columns:
            pct = self.df[col].astype(float).values
        elif '39Ar(k)(%)' in self.df.columns:
            pct = self.df['39Ar(k)(%)'].astype(float).values
        else:
            pct = np.full(len(self.df), 100.0 / max(len(self.df), 1))
        return np.cumsum(pct)

    def _compute_clk(self):
        """Compute Cl/K = 38Ar(Cl)/39Ar(K) × 0.22 with quadrature error."""
        df = self.df
        if '38Ar(cl)' not in df.columns or '39Ar(k)' not in df.columns:
            return None, None
        cl = df['38Ar(cl)'].astype(float)
        k = df['39Ar(k)'].astype(float)
        clk = (cl / k * 0.22).values
        if '38Ar(cl)_std' in df.columns and '39Ar(k)_std' in df.columns:
            cls = df['38Ar(cl)_std'].astype(float).abs()
            ks = df['39Ar(k)_std'].astype(float).abs()
            cl_safe = cl.abs().replace(0, np.nan)
            k_safe = k.abs().replace(0, np.nan)
            cl_rel = (cls / cl_safe).fillna(0).values
            k_rel = (ks / k_safe).fillna(0).values
            clk_std = (np.abs(clk) * np.sqrt(cl_rel**2 + k_rel**2))
        else:
            clk_std = np.zeros(len(df))
        return clk, clk_std

    # ------------------------------------------------------------------
    def _fill_block(self, ws, columns, n_cols):
        """Generic block writer: each row gets `n_cols` numeric values
        from `columns` (a sequence of (length-N) arrays)."""
        n = min(len(columns[0]), TEMPLATE_MAX_STEPS)
        for i in range(n):
            for j in range(n_cols):
                v = columns[j][i]
                ws.cell(row=2 + i, column=j + 1,
                        value=float(v) if pd.notna(v) and not (isinstance(v, float) and np.isnan(v)) else None)
        # Blank tail rows so old template values don't leak
        for i in range(n, TEMPLATE_MAX_STEPS):
            for j in range(n_cols):
                ws.cell(row=2 + i, column=j + 1, value=None)

    def _fill_normal_isochron(self, ws):
        df = self.df
        cols = ['39Ar(m)/36Ar(m)', '39Ar(m)/36Ar(m)_std',
                '40Ar(m)/36Ar(m)', '40Ar(m)/36Ar(m)_std']
        if not all(c in df.columns for c in cols):
            print("[ExcelExport] normal isochron columns missing — skipping")
            return
        arrs = [df[c].astype(float).values for c in cols]
        self._fill_block(ws, arrs, 4)

    def _fill_inverse_isochron(self, ws):
        df = self.df
        cols = ['39Ar(m)/40Ar(m)', '39Ar(m)/40Ar(m)_std',
                '36Ar(m)/40Ar(m)', '36Ar(m)/40Ar(m)_std']
        if not all(c in df.columns for c in cols):
            print("[ExcelExport] inverse isochron columns missing — skipping")
            return
        arrs = [df[c].astype(float).values for c in cols]
        self._fill_block(ws, arrs, 4)

    def _fill_age_spectrum(self, ws, cum):
        df = self.df
        if 'Age(Ma)' not in df.columns:
            print("[ExcelExport] Age(Ma) column missing — skipping")
            return
        age = df['Age(Ma)'].astype(float).values
        sigma = df['Age_std(Ma)'].astype(float).values if 'Age_std(Ma)' in df.columns else np.zeros(len(df))
        self._fill_block(ws, [age, sigma, cum], 3)

    def _fill_cak(self, ws, cum):
        df = self.df
        if 'Ca/K' not in df.columns:
            print("[ExcelExport] Ca/K column missing — skipping")
            return
        cak = df['Ca/K'].astype(float).values
        cak_s = df['Ca/K_std'].astype(float).values if 'Ca/K_std' in df.columns else np.zeros(len(df))
        self._fill_block(ws, [cak, cak_s, cum], 3)

    def _fill_clk(self, ws, cum):
        clk, clk_std = self._compute_clk()
        if clk is None:
            print("[ExcelExport] 38Ar(cl)/39Ar(k) columns missing — skipping Cl/K")
            return
        self._fill_block(ws, [clk, clk_std, cum], 3)

    def _set_sample_id(self, wb):
        try:
            ws = wb['output data']
            sid = self.sample_info.get('Samp#', '')
            mineral = self.sample_info.get('Min', '')
            ws['A1'] = f"Sample ID = {sid}  ({mineral})"
        except Exception:
            pass

    def _add_summary_sheet(self, wb):
        if 'Summary' in wb.sheetnames:
            del wb['Summary']
        ws = wb.create_sheet('Summary', 0)
        ws['A1'] = 'pyADR Step Heating Summary'
        ws['A1'].font = Font(size=14, bold=True)
        ws.merge_cells('A1:D1')

        r = 3
        ws.cell(row=r, column=1, value='Sample Information').font = Font(bold=True, size=12)
        r += 1
        for k, v in self.sample_info.items():
            ws.cell(row=r, column=1, value=k).font = Font(bold=True)
            try:
                if isinstance(v, float):
                    ws.cell(row=r, column=2, value=v)
                else:
                    ws.cell(row=r, column=2, value=str(v))
            except Exception:
                ws.cell(row=r, column=2, value=str(v))
            r += 1

        r += 1
        ws.cell(row=r, column=1, value='Statistics').font = Font(bold=True, size=12)
        r += 1
        sh = self.stats.get('sh_result') or []
        df_res = self.stats.get('DF_result') or []
        rows = []
        if len(sh) >= 1: rows.append(('Weighted Plateau Age (Ma)', sh[0]))
        if len(sh) >= 2: rows.append(('Total Fusion Age (Ma)', sh[1]))
        labels = [
            ('39Ar/36Ar Intercept (Normal)', 0),
            ('39Ar/36Ar Intercept Std', 1),
            ('36Ar/40Ar Intercept (Inverse)', 2),
            ('36Ar/40Ar Intercept Std', 3),
            ('MSWD', 4),
            ('WMA (Weighted Mean Age, Ma)', 5),
            ('Isochron Age (Ma)', 6),
            ('Isochron Age Std (Ma)', 7),
        ]
        for label, idx in labels:
            if idx < len(df_res):
                rows.append((label, df_res[idx]))
        for label, value in rows:
            ws.cell(row=r, column=1, value=label).font = Font(bold=True)
            try:
                ws.cell(row=r, column=2, value=float(value))
            except Exception:
                ws.cell(row=r, column=2, value=str(value))
            r += 1

        r += 1
        ws.cell(row=r, column=1, value='Step Mask').font = Font(bold=True, size=12)
        r += 1
        ws.cell(row=r, column=1, value='Total steps').font = Font(bold=True)
        ws.cell(row=r, column=2, value=int(self.mask.size))
        r += 1
        ws.cell(row=r, column=1, value='Included steps').font = Font(bold=True)
        ws.cell(row=r, column=2, value=int((self.mask > 0.5).sum()))

        ws.column_dimensions['A'].width = 32
        ws.column_dimensions['B'].width = 22

    # ------------------------------------------------------------------
    def _export_template(self):
        print(f"[ExcelExport] Using TEMPLATE: {self.template_path}")
        os.makedirs(os.path.dirname(self.out_filename) or '.', exist_ok=True)
        shutil.copyfile(self.template_path, self.out_filename)

        wb = load_workbook(self.out_filename)
        cum = self._cumulative_percent_array()

        # Map of sheet name → fill function
        # Note: sheet names may have hyphen variants. Match flexibly.
        sheet_fillers = {
            'normal isochron':  lambda ws: self._fill_normal_isochron(ws),
            'inverse isochron': lambda ws: self._fill_inverse_isochron(ws),
            'Age spectrum':     lambda ws: self._fill_age_spectrum(ws, cum),
            'Ca-K ratio':       lambda ws: self._fill_cak(ws, cum),
            'Cl-K ratio':       lambda ws: self._fill_clk(ws, cum),
            # Backward-compat alias (older template used 'Ca K ratio')
            'Ca K ratio':       lambda ws: self._fill_cak(ws, cum),
            'Cl K ratio':       lambda ws: self._fill_clk(ws, cum),
        }
        for name, fn in sheet_fillers.items():
            if name in wb.sheetnames:
                fn(wb[name])
                print(f"[ExcelExport] OK {name} filled")

        self._set_sample_id(wb)
        self._add_summary_sheet(wb)

        wb.save(self.out_filename)
        print(f"[ExcelExport] OK saved: {self.out_filename}")
        return self.out_filename

    # ------------------------------------------------------------------
    def _export_native_fallback(self):
        print("[ExcelExport] Template not found — fallback")
        wb = Workbook()
        ws = wb.active
        ws.title = 'Summary'
        ws['A1'] = 'pyADR Excel export (template missing)'
        ws['A2'] = f"Looked for {DEFAULT_TEMPLATE} next to ExcelChartExporter.py"
        ws['A3'] = "Place the template file there to get the formatted charts."

        ws2 = wb.create_sheet('raw_data')
        for j, c in enumerate(self.df.columns, 1):
            ws2.cell(row=1, column=j, value=str(c))
        for i in range(len(self.df)):
            for j, c in enumerate(self.df.columns, 1):
                v = self.df[c].iloc[i]
                if pd.isna(v):
                    v = ''
                ws2.cell(row=2 + i, column=j, value=v)

        os.makedirs(os.path.dirname(self.out_filename) or '.', exist_ok=True)
        wb.save(self.out_filename)
        print(f"[ExcelExport] OK saved (fallback): {self.out_filename}")
        return self.out_filename

    # ------------------------------------------------------------------
    def export(self, diagrams=None, **kwargs):
        # diagrams arg ignored — template defines what gets exported
        self.read_data()
        if self.template_path:
            return self._export_template()
        return self._export_native_fallback()


def export_diagrams_to_excel(data_file, mask, constants,
                             out_filename='pyADR_Export.xlsx',
                             diagrams=None,
                             work_dir=None, stats=None, template_path=None):
    exporter = ExcelChartExporter(data_file, mask, constants, out_filename,
                                  work_dir=work_dir, stats=stats,
                                  template_path=template_path)
    return exporter.export(diagrams=diagrams)


if __name__ == '__main__':
    print("ExcelChartExporter (clean template edition) loaded successfully")
